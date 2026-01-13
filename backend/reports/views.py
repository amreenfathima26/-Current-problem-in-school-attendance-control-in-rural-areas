from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Count, Q
from attendance.models import AttendanceRecord
from users.models import Student, User, School
from datetime import datetime, date
import csv
import logging
import json
from io import BytesIO

logger = logging.getLogger(__name__)

# Try to import reportlab for PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logger.warning("ReportLab not installed. PDF generation will fail.")

# Try to import openpyxl for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("OpenPyXL not installed. Excel generation will fail.")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_daily_report(request, date):
    """Generate PDF report for daily attendance."""
    if not HAS_REPORTLAB:
        return Response({'error': 'PDF libraries not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    try:
        # Parse date
        report_date = datetime.strptime(date, '%Y-%m-%d').date()
        
        # Get attendance records
        records = AttendanceRecord.objects.filter(date=report_date).select_related('student', 'student__user')
        
        # Calculate stats
        total_students = Student.objects.filter(is_active=True).count()
        present_count = records.filter(status='present').count()
        absent_count = records.filter(status='absent').count()
        late_count = records.filter(status='late').count()
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        elements.append(Paragraph(f"Daily Attendance Report: {report_date}", title_style))
        elements.append(Spacer(1, 12))
        
        # Summary Table
        summary_data = [
            ['Total Students', 'Present', 'Absent', 'Late', 'Attendance %'],
            [
                total_students, 
                present_count, 
                absent_count, 
                late_count, 
                f"{((present_count + late_count)/total_students*100) if total_students > 0 else 0:.1f}%"
            ]
        ]
        
        t = Table(summary_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 24))
        
        # Detail Table
        data = [['Student ID', 'Name', 'Grade', 'Status', 'Time', 'Method']]
        
        for record in records:
            data.append([
                record.student.student_id,
                record.student.user.get_full_name(),
                record.student.grade,
                record.status.upper(),
                record.timestamp.strftime('%H:%M:%S') if record.timestamp else '-',
                record.method
            ])
            
        t2 = Table(data)
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white])
        ]))
        elements.append(t2)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="daily_report_{date}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Error generating daily report: {e}", exc_info=True)
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_monthly_report(request, year, month):
    """Generate PDF report for monthly attendance."""
    if not HAS_REPORTLAB:
        return Response({'error': 'PDF libraries not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    try:
        # Convert month name/number to number
        try:
            month_num = datetime.strptime(month, '%B').month
        except ValueError:
             try:
                 month_num = int(month)
             except ValueError:
                 return Response({'error': 'Invalid month format'}, status=status.HTTP_400_BAD_REQUEST)

        start_date = date(year, month_num, 1)
        if month_num == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month_num + 1, 1)
            
        records = AttendanceRecord.objects.filter(
            date__gte=start_date,
            date__lt=end_date
        ).select_related('student', 'student__user')
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Monthly Attendance Report: {month} {year}", styles['Title']))
        elements.append(Spacer(1, 20))
        
        # Aggregate by student
        student_stats = {}
        all_students = Student.objects.filter(is_active=True).select_related('user')
        
        for student in all_students:
            student_stats[student.id] = {
                'name': student.user.get_full_name(),
                'id': student.student_id,
                'grade': student.grade,
                'present': 0,
                'absent': 0,
                'late': 0,
                'excused': 0
            }
            
        for record in records:
            if record.student.id in student_stats:
                student_stats[record.student.id][record.status] += 1
                
        # Table
        data = [['Student ID', 'Name', 'Grade', 'Present', 'Absent', 'Late', 'Excused', 'Attendance %']]
        
        # Assume 20 working days for percentage calc, or calculate actual
        total_days = 22 
        
        for sid, stats in student_stats.items():
            total_attended = stats.get('present', 0) + stats.get('late', 0)
            percentage = (total_attended / total_days * 100) if total_days > 0 else 0
            percentage = min(percentage, 100.0) # Cap at 100
            
            data.append([
                stats['id'],
                stats['name'],
                stats['grade'],
                stats.get('present', 0),
                stats.get('absent', 0),
                stats.get('late', 0),
                stats.get('excused', 0),
                f"{percentage:.1f}%"
            ])
            
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="monthly_report_{year}_{month}.pdf"'
        return response

    except Exception as e:
        logger.error(f"Error generating monthly report: {e}", exc_info=True)
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_student_report(request, student_id):
    """Generate PDF report for specific student history."""
    if not HAS_REPORTLAB:
        return Response({'error': 'PDF libraries not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    try:
        student = Student.objects.get(student_id=student_id)
        records = AttendanceRecord.objects.filter(student=student).order_by('-date')
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Student Attendance Report: {student.user.get_full_name()}", styles['Title']))
        elements.append(Paragraph(f"ID: {student.student_id} | Grade: {student.grade}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        data = [['Date', 'Status', 'Time', 'Method', 'Notes']]
        
        for record in records:
            data.append([
                record.date.strftime('%Y-%m-%d'),
                record.status.upper(),
                record.timestamp.strftime('%H:%M:%S') if record.timestamp else '-',
                record.method,
                record.notes or ''
            ])
            
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="student_report_{student_id}.pdf"'
        return response

    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error generating student report: {e}", exc_info=True)
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_excel(request):
    """Export attendance to Excel."""
    if not HAS_OPENPYXL:
         return Response({'error': 'Excel libraries not installed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        records = AttendanceRecord.objects.all().select_related('student', 'student__user').order_by('date')
        
        if start_date_str:
            records = records.filter(date__gte=start_date_str)
        if end_date_str:
             records = records.filter(date__lte=end_date_str)
             
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"
        
        # Headers
        headers = ['Date', 'Student ID', 'Name', 'Grade', 'Status', 'Time', 'Method', 'Notes']
        for col_num, header in enumerate(headers, 1):
             cell = ws.cell(row=1, column=col_num, value=header)
             cell.font = Font(bold=True)
             
        # Data
        for row_num, record in enumerate(records, 2):
             ws.cell(row=row_num, column=1, value=record.date)
             ws.cell(row=row_num, column=2, value=record.student.student_id)
             ws.cell(row=row_num, column=3, value=record.student.user.get_full_name())
             ws.cell(row=row_num, column=4, value=record.student.grade)
             ws.cell(row=row_num, column=5, value=record.status)
             ws.cell(row=row_num, column=6, value=record.timestamp.strftime('%H:%M:%S') if record.timestamp else '')
             ws.cell(row=row_num, column=7, value=record.method)
             ws.cell(row=row_num, column=8, value=record.notes)
             
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="attendance_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Error exporting excel: {e}", exc_info=True)
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
