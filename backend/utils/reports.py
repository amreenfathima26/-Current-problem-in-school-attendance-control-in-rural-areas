"""
Report generation utilities for EDURFID system.
"""
# import pandas as pd  # Commented out - not compatible with Python 3.13
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus import PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import io
import logging

logger = logging.getLogger(__name__)


class AttendanceReportGenerator:
    """Generator for attendance reports in various formats."""
    
    def __init__(self):
        """Initialize report generator."""
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()

    def setup_custom_styles(self):
        """Setup custom styles for reports."""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        ))
        
        # Heading style
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.darkblue
        ))
        
        # Normal style
        self.styles.add(ParagraphStyle(
            name='CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        ))

    def generate_daily_attendance_pdf(self, date: datetime, attendance_data: Dict[str, Any]) -> bytes:
        """
        Generate daily attendance report in PDF format.
        
        Args:
            date: Date for the report
            attendance_data: Dictionary containing attendance data
            
        Returns:
            bytes: PDF content
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        story = []

        # Title
        title = Paragraph(f"Daily Attendance Report - {date.strftime('%B %d, %Y')}", 
                         self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))

        # Summary section
        summary_data = [
            ['Total Students', str(attendance_data.get('total_students', 0))],
            ['Present', str(attendance_data.get('present_count', 0))],
            ['Absent', str(attendance_data.get('absent_count', 0))],
            ['Late', str(attendance_data.get('late_count', 0))],
            ['Excused', str(attendance_data.get('excused_count', 0))],
            ['Attendance %', f"{attendance_data.get('attendance_percentage', 0):.1f}%"]
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(Paragraph("Attendance Summary", self.styles['CustomHeading']))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Student details section
        if attendance_data.get('students'):
            story.append(Paragraph("Student Details", self.styles['CustomHeading']))
            
            # Prepare student data
            student_data = [['Student Name', 'Grade', 'Status', 'Time']]
            for student in attendance_data['students']:
                student_data.append([
                    student.get('student_name', ''),
                    student.get('grade', ''),
                    student.get('status', ''),
                    student.get('timestamp', '')[:19] if student.get('timestamp') else ''
                ])

            student_table = Table(student_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1.5*inch])
            student_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))

            story.append(student_table)

        # Footer
        story.append(Spacer(1, 30))
        footer = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                          self.styles['CustomNormal'])
        story.append(footer)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_monthly_attendance_pdf(self, month: int, year: int, 
                                       attendance_data: List[Dict[str, Any]]) -> bytes:
        """
        Generate monthly attendance report in PDF format.
        
        Args:
            month: Month number (1-12)
            year: Year
            attendance_data: List of daily attendance data
            
        Returns:
            bytes: PDF content
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        story = []

        # Title
        month_name = datetime(year, month, 1).strftime('%B')
        title = Paragraph(f"Monthly Attendance Report - {month_name} {year}", 
                         self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))

        # Summary statistics
        total_days = len(attendance_data)
        avg_attendance = sum(day.get('attendance_percentage', 0) for day in attendance_data) / total_days if total_days > 0 else 0
        best_day = max(attendance_data, key=lambda x: x.get('attendance_percentage', 0)) if attendance_data else {}
        worst_day = min(attendance_data, key=lambda x: x.get('attendance_percentage', 0)) if attendance_data else {}

        summary_data = [
            ['Total Days', str(total_days)],
            ['Average Attendance %', f"{avg_attendance:.1f}%"],
            ['Best Day', f"{best_day.get('date', '')} ({best_day.get('attendance_percentage', 0):.1f}%)"],
            ['Worst Day', f"{worst_day.get('date', '')} ({worst_day.get('attendance_percentage', 0):.1f}%)"]
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE(Model', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(Paragraph("Monthly Summary", self.styles['CustomHeading']))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Daily breakdown
        if attendance_data:
            story.append(Paragraph("Daily Breakdown", self.styles['CustomHeading']))
            
            # Prepare daily data
            daily_data = [['Date', 'Present', 'Absent', 'Late', 'Excused', 'Attendance %']]
            for day in attendance_data:
                daily_data.append([
                    day.get('date', ''),
                    str(day.get('present_count', 0)),
                    str(day.get('absent_count', 0)),
                    str(day.get('late_count', 0)),
                    str(day.get('excused_count', 0)),
                    f"{day.get('attendance_percentage', 0):.1f}%"
                ])

            daily_table = Table(daily_data, colWidths=[1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.2*inch])
            daily_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))

            story.append(daily_table)

        # Footer
        story.append(Spacer(1, 30))
        footer = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                          self.styles['CustomNormal'])
        story.append(footer)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_student_attendance_pdf(self, student_data: Dict[str, Any]) -> bytes:
        """
        Generate individual student attendance report in PDF format.
        
        Args:
            student_data: Dictionary containing student attendance data
            
        Returns:
            bytes: PDF content
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        story = []

        # Title
        student_name = student_data.get('student', {}).get('user', {}).get('full_name', 'Unknown Student')
        title = Paragraph(f"Student Attendance Report - {student_name}", 
                         self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))

        # Student info
        student_info = [
            ['Student Name', student_name],
            ['Student ID', student_data.get('student', {}).get('student_id', '')],
            ['Grade', student_data.get('student', {}).get('grade', '')],
            ['Total Days', str(student_data.get('total_days', 0))],
            ['Present Days', str(student_data.get('present_days', 0))],
            ['Absent Days', str(student_data.get('absent_days', 0))],
            ['Late Days', str(student_data.get('late_days', 0))],
            ['Excused Days', str(student_data.get('excused_days', 0))],
            ['Attendance %', f"{student_data.get('attendance_percentage', 0):.1f}%"]
        ]

        info_table = Table(student_info, colWidths=[3*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(Paragraph("Student Information", self.styles['CustomHeading']))
        story.append(info_table)
        story.append(Spacer(1, 20))

        # Recent attendance records
        if student_data.get('recent_records'):
            story.append(Paragraph("Recent Attendance Records", self.styles['CustomHeading']))
            
            recent_data = [['Date', 'Status', 'Notes']]
            for record in student_data['recent_records'][:20]:  # Limit to last 20 records
                recent_data.append([
                    record.get('date', ''),
                    record.get('status', ''),
                    record.get('notes', '')[:50]  # Truncate long notes
                ])

            recent_table = Table(recent_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
            recent_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))

            story.append(recent_table)

        # Footer
        story.append(Spacer(1, 30))
        footer = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                          self.styles['CustomNormal'])
        story.append(footer)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_attendance_excel(self, attendance_data: List[Dict[str, Any]], 
                                 filename: str = None) -> bytes:
        """
        Generate attendance report in Excel format.
        
        Args:
            attendance_data: List of attendance data
            filename: Optional filename
            
        Returns:
            bytes: Excel content
        """
        # Simplified Excel generation without pandas
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        if attendance_data:
            # Write headers
            headers = list(attendance_data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, record in enumerate(attendance_data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=record.get(header, ''))
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_summary_statistics(self, attendance_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate summary statistics from attendance data.
        
        Args:
            attendance_data: List of attendance data
            
        Returns:
            Dictionary with summary statistics
        """
        if not attendance_data:
            return {}

        # Calculate statistics without pandas
        total_records = len(attendance_data)
        
        # Calculate averages and sums manually
        attendance_percentages = [item.get('attendance_percentage', 0) for item in attendance_data if 'attendance_percentage' in item]
        present_counts = [item.get('present_count', 0) for item in attendance_data if 'present_count' in item]
        absent_counts = [item.get('absent_count', 0) for item in attendance_data if 'absent_count' in item]
        late_counts = [item.get('late_count', 0) for item in attendance_data if 'late_count' in item]
        excused_counts = [item.get('excused_count', 0) for item in attendance_data if 'excused_count' in item]
        
        stats = {
            'total_records': total_records,
            'average_attendance': sum(attendance_percentages) / len(attendance_percentages) if attendance_percentages else 0,
            'highest_attendance': max(attendance_percentages) if attendance_percentages else 0,
            'lowest_attendance': min(attendance_percentages) if attendance_percentages else 0,
            'total_present': sum(present_counts),
            'total_absent': sum(absent_counts),
            'total_late': sum(late_counts),
            'total_excused': sum(excused_counts),
        }
        
        return stats


# Example usage
if __name__ == "__main__":
    # Create report generator
    generator = AttendanceReportGenerator()
    
    # Sample data
    sample_data = {
        'date': '2024-01-15',
        'total_students': 50,
        'present_count': 45,
        'absent_count': 3,
        'late_count': 2,
        'excused_count': 0,
        'attendance_percentage': 94.0,
        'students': [
            {'student_name': 'John Doe', 'grade': '10', 'status': 'present', 'timestamp': '2024-01-15T08:30:00'},
            {'student_name': 'Jane Smith', 'grade': '10', 'status': 'present', 'timestamp': '2024-01-15T08:25:00'},
        ]
    }
    
    # Generate PDF report
    pdf_content = generator.generate_daily_attendance_pdf(datetime.now(), sample_data)
    
    # Save to file
    with open('daily_attendance_report.pdf', 'wb') as f:
        f.write(pdf_content)
    
    print("Report generated successfully!")
